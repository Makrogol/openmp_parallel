import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.chart import LineChart, Reference
from openpyxl.styles import Alignment, Font


f = open('input_data/variabels.txt')
count_sizes = 0
max_count_threads = 0
for line in f.readlines():
    if line.find('COUNT_SIZES') != -1:
        count_sizes = int(line.split('=')[-1])
    if line.find('MAX_COUNT_THREADS') != -1:
        max_count_threads = int(line.split('=')[-1])
f.close()

f = open('input_data/sizes.txt', 'r')
sizes = [line.strip() for line in f.readlines()]
f.close()

colors = ['FF0000', '000000', '0000FF', '00FF00']
solution_types_str = ['Параллельное решение',
                      'Параллельное решение с reduction', 
                      'Параллельное решение с section']


wb = openpyxl.load_workbook('out/results.xlsx')
sheet = wb.active
sheet.delete_rows(4)

# unmerging
max_column = sheet.max_column
sheet.unmerge_cells(start_row=1, start_column=2,
                    end_row=1, end_column=max_column)

for i in range(count_sizes, 0, -1):
    end_column = i * 4 + 1
    start_column = end_column - 3
    sheet.unmerge_cells(start_row=2, start_column=start_column,
                        end_row=2, end_column=end_column)
    if i > 1:
        sheet.insert_cols(start_column)
        sheet.insert_cols(start_column)
        for row_index in range(3, sheet.max_row + 1):
            sheet.cell(row=row_index, column=start_column + 1,
                       value=sheet.cell(row=row_index, column=1).value)
# +51
# время
for column_index in range(2, sheet.max_column, 6):
    for i in range(4):
        sheet.column_dimensions[get_column_letter(column_index + i)].width = 35

    sheet.merge_cells(start_row=2, start_column=column_index,
                      end_row=2, end_column=column_index + 3)

    data = Reference(sheet, min_col=column_index, min_row=3,
                     max_col=column_index + 3, max_row=sheet.max_row)
    chart = LineChart()
    chart.title = sheet.cell(row=2, column=column_index).value
    chart.x_axis.title = 'Количество потоков'
    chart.width = 30
    chart.height = 20
    chart.add_data(data, titles_from_data=True)

    for i in range(len(chart.series)):
        chart.series[i].graphicalProperties.line.solidFill = colors[i]
        chart.series[i].marker.symbol = 'circle'
        chart.series[i].marker.size = 8
        chart.series[i].marker.graphicalProperties.line.solidFill = colors[i]

    sheet.add_chart(chart, get_column_letter(
        column_index - 1) + str(sheet.max_row + 2))

sheet.cell(row=1, column=2, value='Время выполнения')
sheet.merge_cells(start_row=1, start_column=2,
                  end_row=1, end_column=sheet.max_column)

# Ускорение
for column_index in range(1, sheet.max_column, 6):
    for row in range(2, 2 + max_count_threads + 2):
        sheet.cell(row=row + 51, column=column_index, value=sheet.cell(row=row, column=column_index).value)
        sheet.cell(row=row + 51, column=column_index).font = Font(bold=True)

    sheet.cell(row=2 + 51, column=column_index + 1, value=sizes[column_index // 6])

    for col in range(column_index + 2, column_index + 5):
        for row in range(2, 2 + 2):
            sheet.cell(row=row + 51, column=col - 1, value=sheet.cell(row=row, column=col).value)
            sheet.cell(row=row + 51, column=col - 1).font = Font(bold=True)

        for row in range(4, 4 + max_count_threads):
            sequential_val = float(sheet.cell(row=row, column=column_index + 1).value)
            parallel_val = float(sheet.cell(row=row, column=col).value)
            sheet.cell(row=row + 51, column=col - 1, value=round(sequential_val / parallel_val, 2))

sheet.cell(row=1 + 51, column=2, value='Ускорение')
sheet.merge_cells(start_row=1 + 51, start_column=2,
                  end_row=1 + 51, end_column=sheet.max_column)

for column_index in range(2, sheet.max_column, 6):
    sheet.merge_cells(start_row=2 + 51, start_column=column_index,
                      end_row=2 + 51, end_column=column_index + 3)

    data = Reference(sheet, min_col=column_index, min_row=3 + 51,
                    max_col=column_index + 2, max_row=sheet.max_row)
    chart = LineChart()
    chart.title = sheet.cell(row=2 + 51, column=column_index).value
    chart.x_axis.title = 'Количество потоков'
    chart.width = 30
    chart.height = 20
    chart.add_data(data, titles_from_data=True)

    for i in range(len(chart.series)):
        chart.series[i].graphicalProperties.line.solidFill = colors[i]
        chart.series[i].marker.symbol = 'circle'
        chart.series[i].marker.size = 8
        chart.series[i].marker.graphicalProperties.line.solidFill = colors[i]

    sheet.add_chart(chart, get_column_letter(
        column_index - 1) + str(sheet.max_row + 2))
    


# Эффективность
for column_index in range(1, sheet.max_column, 6):
    for row in range(2 + 51, 2 + 51 + max_count_threads + 2):
        sheet.cell(row=row + 51, column=column_index, value=sheet.cell(row=row, column=column_index).value)
        sheet.cell(row=row + 51, column=column_index).font = Font(bold=True)

    sheet.cell(row=2 + 51 * 2, column=column_index + 1, value=sizes[column_index // 6])

    for col in range(column_index + 1, column_index + 4):
        for row in range(2 + 51, 2 + 2 + 51):
            sheet.cell(row=row + 51, column=col, value=sheet.cell(row=row, column=col).value)
            sheet.cell(row=row + 51, column=col).font = Font(bold=True)

        for row in range(4 + 51, 4 + max_count_threads + 51):
            acceleration_val = float(sheet.cell(row=row, column=col).value)
            count_threads = float(sheet.cell(row=row, column=column_index).value)
            sheet.cell(row=row + 51, column=col, value=round(acceleration_val / count_threads, 4) * 100)

sheet.cell(row=1 + 51 * 2, column=2, value='Эффективность')
sheet.merge_cells(start_row=1 + 51 * 2, start_column=2,
                  end_row=1 + 51 * 2, end_column=sheet.max_column)

for column_index in range(2, sheet.max_column, 6):
    sheet.merge_cells(start_row=2 + 51 * 2, start_column=column_index,
                      end_row=2 + 51 * 2, end_column=column_index + 3)

    data = Reference(sheet, min_col=column_index, min_row=3 + 51 * 2,
                    max_col=column_index + 2, max_row=sheet.max_row)
    chart = LineChart()
    chart.title = sheet.cell(row=2 + 51 * 2, column=column_index).value
    chart.x_axis.title = 'Количество потоков'
    chart.width = 30
    chart.height = 20
    chart.add_data(data, titles_from_data=True)

    for i in range(len(chart.series)):
        chart.series[i].graphicalProperties.line.solidFill = colors[i]
        chart.series[i].marker.symbol = 'circle'
        chart.series[i].marker.size = 8
        chart.series[i].marker.graphicalProperties.line.solidFill = colors[i]

    sheet.add_chart(chart, get_column_letter(
        column_index - 1) + str(sheet.max_row + 2))


for col in range(1, sheet.max_column + 1):
    for row in range(1, sheet.max_row + 1):
        sheet.cell(row=row, column=col).alignment = Alignment(horizontal='center')

wb.save('out/results.xlsx')
